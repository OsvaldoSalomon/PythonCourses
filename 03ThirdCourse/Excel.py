# Workbooks
# Creating a new workbook
import openpyxl

# workbook = Workbook() //Creating a new workbook object, by initializing the Workbook() class

workbook = openpyxl.load_workbook("E:/osvaldo/PythonExercises/03ThirdCourse/Excel/Employees.xlsx")

# Getting the basic properties of the workbook
print(workbook.properties)

# Getting the sheets in the workbook
print(workbook.worksheets)

# Getting the sheet names in a workbook, as a list of strings
print(workbook.sheetnames)

# Getting the active sheet in the workbook
print(workbook.active)

print(" --------- ")

# Referencing a sheet by its name
sheet = workbook["EmployeeData"]

# Creating a new sheet in the workbook
# print(workbook.create_sheet('TestSheet'))

# Removing a sheet from the workbook
# print(workbook.remove(sheet))

print("Active cell:", sheet.active_cell)

print("Dimensions:", sheet.dimensions)

print("Max Rows:", sheet.max_row)
print("Max Columns:", sheet.max_column)

print(" --------- ")

print(sheet['B7'].value)
print(sheet.cell(row=6, column=2).value)

print(" --------- ")

cell = sheet['B9']
print(cell.row)
print(cell.column)
print(cell.coordinate)
print(cell.data_type)

print(" --------- ")

cell = sheet['B2']
cell.value = 'David'


# Saving the workbook
workbook.save("E:/osvaldo/PythonExercises/03ThirdCourse/Excel/Employees.xlsx")

# workbook.close()
